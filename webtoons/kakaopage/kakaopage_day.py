from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import openpyxl
import time

chrome_driver = "C:\\Users\\s6xya\\Desktop\\selenium\\chromedriver.exe"
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:8989")
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

# *day=? mon:1 ~ sun:7, all:12
driver.get('https://page.kakao.com/main?categoryUid=10&subCategoryUid=1002&day=12')
driver.implicitly_wait(5)
print("kakaopage_day")


# 완결 장르 소년/드라마/로맨스/로판/액션무협/BL&GL
# time.sleep(1)
# driver.find_elements_by_class_name("css-1tpompv")[0].click()
# driver.find_elements_by_css_selector(".css-1lx7oj")[1].click()
# time.sleep(1)

#####################################################
WB = openpyxl.Workbook()  # 엑셀 파일 생성
WB.create_sheet("kakaopage")  # 시트 추가
sheet = WB["kakaopage"]  # 시트 선택
sheet['A1'] = "id"
sheet["B1"] = "title"
sheet["C1"] = "author"
sheet["D1"] = "synopsis"
sheet["E1"] = "genre"
sheet["F1"] = "age"
sheet["G1"] = "likes"
sheet["H1"] = "image"
sheet["I1"] = "url"
sheet["J1"] = "day"
#####################################################


def scrollDown():
    scrolled = 0
    SCROLL_PAUSE_TIME = 1
    # *change max_height
    max_height = 5569
    last_height = driver.execute_script("return document.body.scrollHeight")

    while last_height <= max_height:
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE_TIME)
        new_height = driver.execute_script("return document.body.scrollHeight")
        print("scrolling..." + str(scrolled))
        scrolled = scrolled + 1
        if new_height == last_height:
            scrolled = 0
            break
        last_height = new_height

#####################################################


def addWebtoon():
    count = 0
    indexError = []
    webtoons = driver.find_elements_by_css_selector('.css-19y0ur2 a')

    # test
    # for z in [0, 1, 2, 3]:
    # for i in range(0, 933):
    for webtoon in webtoons:
        try:
            x = driver.find_elements_by_css_selector(
                '.css-19y0ur2 a')[count].get_attribute('href')
            driver.get(x)
            driver.implicitly_wait(5)

            title = driver.find_elements_by_css_selector(".css-jgjrt")[0].text
            author = driver.find_elements_by_css_selector(
                ".css-7a7cma")[1].text
            day = driver.find_elements_by_css_selector(".css-7a7cma")[0].text
            likes = driver.find_elements_by_css_selector(".css-fjzm5m")[0].text
            image = driver.find_elements_by_css_selector(
                ".css-1ithwm4")[0].get_attribute("src")

            # 상세페이지 접속
            detail = driver.find_elements_by_css_selector(
                ".css-zkp4tp")[0].send_keys(Keys.ENTER)

            synopsis = driver.find_elements_by_css_selector(
                ".contentCol .descriptionBox")[0].text
            genre = driver.find_elements_by_css_selector(
                ".contentCol .infoBox .jsx-3755015728 .css-1l2pl53")[0].text
            age = driver.find_elements_by_css_selector(
                ".contentCol .infoBox")[2].text

            url = driver.current_url

            sheet.cell(count+2, 1, count)
            sheet.cell(count+2, 2, title)
            sheet.cell(count+2, 3, author)
            sheet.cell(count+2, 4, synopsis)
            sheet.cell(count+2, 5, genre)
            sheet.cell(count+2, 6, age)
            sheet.cell(count+2, 7, likes)
            sheet.cell(count+2, 8, image)
            sheet.cell(count+2, 9, url)
            sheet.cell(count+2, 10, day)

            count = count + 1
            print(title)
            print(count)
            print(len(webtoons))

            driver.back()
            scrollDown()
            driver.implicitly_wait(5)
        except IndexError:
            age = "null"
            print("-----error-----")
            indexError.append(count)
            print(indexError)
            print("-----error-----")
            count = count + 1

    WB.save("kakopage.xls")
    print("DONE!")

#####################################################


scrollDown()
time.sleep(5)
addWebtoon()
